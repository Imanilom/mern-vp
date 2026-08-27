import json, datetime, sys, os
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# Read .env for MONGO uri
mongo_uri = 'mongodb://127.0.0.1:27017/'
env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
if os.path.exists(env_path):
    with open(env_path, 'r') as f:
        for line in f:
            if line.startswith('MONGO='):
                mongo_uri = line.split('=', 1)[1].strip().strip('"').strip("'")
                break

# Connect to MongoDB
try:
    client = MongoClient(mongo_uri, serverSelectionTimeoutMS=5000)
    db = client.get_default_database() if client.get_default_database().name else client['capar-vp']
    client.server_info() # trigger connection check
except Exception as e:
    print("Gagal terkoneksi ke database MongoDB lokal:", e)
    sys.exit(1)

# Find an event to process. Look for PERSISTENT_DEVIATION or any long episode
print("Mencari episode persisten atau anomali durasi terpanjang...")
event = None
episodeDoc = None

eps = list(db.episodeanalyses.find({"end_time": {"$exists": True, "$ne": None}}).sort([("total_duration", -1)]).limit(1))
if len(eps) > 0:
    episodeDoc = eps[0]
    if "episode_id" in episodeDoc:
        event = db.anomalievents.find_one({"_id": episodeDoc["episode_id"]})

if not event:
    events = list(db.anomalievents.find({"status": {"$in": ["closed", "resolved", "recovered"]}}).sort([("duration_ms", -1)]).limit(1))
    if len(events) > 0:
        event = events[0]
        episodeDoc = db.episodeanalyses.find_one({"episode_id": event["_id"]})

if not event and not episodeDoc:
    print("Tidak ditemukan data anomaly event atau episode di database.")
    sys.exit(1)

# Jika salah satu tidak ada, fallback ke dictionary kosong agar tidak error
if not event:
    event = {}
if not episodeDoc:
    episodeDoc = {"tau_normal": 0.75}

event_id = event.get('_id', episodeDoc.get('_id', 'UNKNOWN'))
print(f"Menggunakan Event/Episode ID: {event_id}")

uid = event.get('user_id', episodeDoc.get('user_id', 'UNKNOWN'))
onset = event.get('onset_time')
if not onset:
    onset = episodeDoc.get('start_time')
    if onset and isinstance(onset, datetime.datetime):
        onset = int(onset.timestamp() * 1000)
    elif onset and isinstance(onset, str):
        # rough parse
        pass

if not onset:
    onset = 0

peak = event.get('peak_time', onset)
resolved = event.get('resolved_time', event.get('recovered_at'))
if not resolved:
    resolved = episodeDoc.get('end_time')
    if resolved and isinstance(resolved, datetime.datetime):
        resolved = int(resolved.timestamp() * 1000)
if not resolved:
    resolved = onset + event.get('duration_ms', episodeDoc.get('total_duration', 15*60000))

# Ambil data dari DB berdasarkan time window +- 30 menit
ctx_start = onset - 1800000
ctx_end = resolved + 1800000

print(f"Mengambil data segment dan raw dari {ctx_start} s/d {ctx_end}...")
segments = list(db.segments.find({"user_id": uid, "window_start": {"$gte": ctx_start, "$lte": ctx_end}}).sort("window_start", 1))
polar = list(db.polardatas.find({"user_id": uid, "timestamp": {"$gte": ctx_start, "$lte": ctx_end}}).sort("timestamp", 1))

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FONT = Font(name=FONT, color="0000FF")           
FORMULA_FONT = Font(name=FONT, color="000000")          
HEADER_FONT = Font(name=FONT, color="FFFFFF", bold=True)
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="1F4E78")
SUB_FONT = Font(name=FONT, size=11, bold=True)
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666666")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def ms_to_dt(ms):
    return datetime.datetime.utcfromtimestamp(ms/1000)

def style_header_row(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

wb = Workbook()
wb.remove(wb.active)

# ============================================================
# OVERVIEW SHEET
# ============================================================
ov = wb.create_sheet("Overview")
ov.sheet_view.showGridLines = False
ov.column_dimensions['A'].width = 3
ov.column_dimensions['B'].width = 34
ov.column_dimensions['C'].width = 90

ov['B2'] = "Analisis Episode Anomali HR — Data Mentah, TTR & AUC-D"
ov['B2'].font = TITLE_FONT
ov['B3'] = f"User ID: {uid}  |  Device: POLAR_DUMP"
ov['B3'].font = Font(name=FONT, italic=True, color="555555")

r = 5
ov.cell(row=r, column=2, value="Sumber data").font = SUB_FONT
r += 1
sources = [
    ("anomalievents", "Metadata episode anomali (onset/peak/resolved, durasi, jeda/pause)."),
    ("episodeanalyses", "Hasil analisis episode (skor, evaluasi model E1-E6, tau threshold)."),
    ("segments", "Data tersegmentasi per jendela 5 menit (anomaly_score, klasifikasi, fitur HRV)."),
    ("polardatas", "Data mentah perangkat Polar (HR, RR, ECG, akselerasi) per sampel ~2 detik."),
]
for name, desc in sources:
    ov.cell(row=r, column=2, value=name).font = Font(name=FONT, bold=True, size=9)
    ov.cell(row=r, column=3, value=desc).font = Font(name=FONT, size=9)
    r += 1

r += 1
ov.cell(row=r, column=2, value="Catatan penting").font = SUB_FONT
r += 1
notes = [
    "Data dirender langsung menggunakan PyMongo dari MongoDB lokal sesuai kriteria persisten/durasi tertinggi.",
    "Field ttr_min, deviation_auc, dan auc_score pada data sumber bila null/0 dihitung ulang dengan formula excel.",
    "TTR (Time To Recovery) dihitung 2 cara: (a) TTR mentah = resolved_time - peak_time (wall-clock), dan (b) TTR aktif = durasi dikurangi total_paused_ms.",
    "AUC-D (Area Under the Curve - Deviation) dihitung dengan aturan trapesium.",
    "Grafik episode penuh memakai data HR mentah (polardatas)."
]
for note in notes:
    ov.cell(row=r, column=2, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ov.cell(row=r, column=2).font = Font(name=FONT, size=9)
    ov.row_dimensions[r].height = 45
    r += 1

print("Overview sheet done")

# ============================================================
# EVENT SHEET
# ============================================================
label = f"Event 1"
scored_segs = [s for s in segments if s.get('anomaly_score') is not None]
none_segs = [s for s in segments if s.get('anomaly_score') is None]

ws = wb.create_sheet(f"Summary")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 26
ws.column_dimensions['D'].width = 55

ws['B2'] = f"{label} — {event.get('classification', 'UNKNOWN')} ({event.get('status', 'UNKNOWN')})"
ws['B2'].font = TITLE_FONT
ws['B3'] = f"Aktivitas: {event.get('activity', 'UNKNOWN')}  |  Device: {event.get('device_id', 'UNKNOWN')}"
ws['B3'].font = Font(name=FONT, italic=True, color="555555")

r = 5
ws.cell(row=r, column=2, value="METADATA EPISODE (mentah)").font = SUB_FONT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1

meta_rows = [
    ("Episode ID", str(event.get('_id', ''))),
    ("Onset time (ms epoch)", onset),
    ("Peak time (ms epoch)", peak),
    ("Resolved time (ms epoch)", resolved),
    ("Duration (ms)", event.get('duration_ms', 0)),
    ("Total paused (ms)", event.get('total_paused_ms', 0)),
    ("Onset score", event.get('onset_score', 0)),
    ("Peak score", event.get('peak_score', 0)),
    ("Classification", event.get('classification', '')),
    ("Status", event.get('status', '')),
    ("Admin status", event.get('admin_status', '')),
    ("Physiological outcome", event.get('physiological_outcome', '')),
    ("Current state", event.get('current_state', '')),
    ("Window count", event.get('window_count', 0)),
    ("Persistence", event.get('trajectory', {}).get('persistence', 0)),
    ("Baseline HR (bpm)", event.get('baseline_hr', 0)),
    ("Peak HR (bpm)", event.get('peak_hr', 0)),
]
first_meta_row = r
for k, v in meta_rows:
    ws.cell(row=r, column=2, value=k).font = Font(name=FONT, bold=True, size=10)
    c = ws.cell(row=r, column=3, value=v)
    c.font = INPUT_FONT
    r += 1

onset_row = first_meta_row + 1
peak_row = first_meta_row + 2
resolved_row = first_meta_row + 3
duration_row = first_meta_row + 4
paused_row = first_meta_row + 5

r += 1
ws.cell(row=r, column=2, value="Onset (UTC)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=f"=C{onset_row}/86400000+DATE(1970,1,1)").number_format = "yyyy-mm-dd hh:mm:ss"
ws.cell(row=r, column=3).font = FORMULA_FONT
r += 1
ws.cell(row=r, column=2, value="Peak (UTC)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=f"=C{peak_row}/86400000+DATE(1970,1,1)").number_format = "yyyy-mm-dd hh:mm:ss"
ws.cell(row=r, column=3).font = FORMULA_FONT
r += 1
ws.cell(row=r, column=2, value="Resolved (UTC)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=f"=C{resolved_row}/86400000+DATE(1970,1,1)").number_format = "yyyy-mm-dd hh:mm:ss"
ws.cell(row=r, column=3).font = FORMULA_FONT

r += 2
ws.cell(row=r, column=2, value="TTR — TIME TO RECOVERY (formula)").font = SUB_FONT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
ws.cell(row=r, column=2, value="TTR mentah / wall-clock (menit)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=f"=(C{resolved_row}-C{peak_row})/60000").font = FORMULA_FONT
ws.cell(row=r, column=3).number_format = "0.00"
ws.cell(row=r, column=4, value="(resolved_time - peak_time) / 60000").font = NOTE_FONT
r += 1
ws.cell(row=r, column=2, value="TTR aktif, exclude pause (menit)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=f"=(C{duration_row}-C{paused_row})/60000").font = FORMULA_FONT
ws.cell(row=r, column=3).number_format = "0.00"
ws.cell(row=r, column=4, value="(duration_ms - total_paused_ms) / 60000").font = NOTE_FONT

r += 2
ws.cell(row=r, column=2, value="AUC-D — AREA UNDER CURVE (DEVIATION)").font = SUB_FONT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
ws.cell(row=r, column=2, value="tau_normal (ambang baseline)").font = Font(name=FONT, size=10)
tau_row = r
ws.cell(row=r, column=3, value=episodeDoc.get('tau_normal', 0.75)).font = INPUT_FONT
ws.cell(row=r, column=4, value="Sumber: episodeanalyses.tau_normal").font = NOTE_FONT
r += 1
ws.cell(row=r, column=2, value="AUC-D total (skor·menit)").font = Font(name=FONT, size=10, bold=True)
auc_row = r
ws.cell(row=r, column=3, value=f"=SUM('Segments'!H:H)").font = Font(name=FONT, bold=True)
ws.cell(row=r, column=3).number_format = "0.000"
ws.cell(row=r, column=4, value="Jumlah kontribusi trapesium dari sheet Segments").font = NOTE_FONT
r += 1
ws.cell(row=r, column=2, value="AUC-D total (skor·jam)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=f"=C{auc_row}/60").font = FORMULA_FONT
ws.cell(row=r, column=3).number_format = "0.0000"
r += 1
ws.cell(row=r, column=2, value="Peak deviation (peak_score - tau_normal)").font = Font(name=FONT, size=10)
peak_score_row = first_meta_row + 7
ws.cell(row=r, column=3, value=f"=C{peak_score_row}-C{tau_row}").font = FORMULA_FONT
ws.cell(row=r, column=3).number_format = "0.000"

r += 2
ws.cell(row=r, column=2, value="RINGKASAN MONITORING").font = SUB_FONT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
r += 1
ws.cell(row=r, column=2, value="Jumlah window tersegmentasi").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=len(scored_segs) + len(none_segs)).font = INPUT_FONT
r += 1
ws.cell(row=r, column=2, value="Window dengan skor valid").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=len(scored_segs)).font = INPUT_FONT
r += 1
ws.cell(row=r, column=2, value="Window jeda/data tidak cukup (None)").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=len(none_segs)).font = INPUT_FONT
r += 1
ws.cell(row=r, column=2, value="Jumlah sampel HR mentah").font = Font(name=FONT, size=10)
ws.cell(row=r, column=3, value=len(polar)).font = INPUT_FONT

# ---------------- Segments sheet ----------------
ws2 = wb.create_sheet(f"Segments")
ws2.sheet_view.showGridLines = False
headers = ["Window start (ms)", "Window end (ms)", "Window start (UTC)", "Klasifikasi",
           "Anomaly score", "Mid-window (menit sejak epoch)", "Deviasi > tau_normal", "Kontribusi AUC (trapesium)",
           "Label chart (text)"]
for c, h in enumerate(headers, start=1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers))
widths = [14, 14, 20, 12, 13, 24, 16, 20, 18]
for c, w in enumerate(widths, start=1):
    ws2.column_dimensions[get_column_letter(c)].width = w

row0 = 2
for i, s in enumerate(scored_segs):
    rr = row0 + i
    ws2.cell(row=rr, column=1, value=s.get('window_start')).font = INPUT_FONT
    ws2.cell(row=rr, column=2, value=s.get('window_end')).font = INPUT_FONT
    c3 = ws2.cell(row=rr, column=3, value=f"=A{rr}/86400000+DATE(1970,1,1)")
    c3.number_format = "yyyy-mm-dd hh:mm:ss"
    c3.font = FORMULA_FONT
    ws2.cell(row=rr, column=4, value=s.get('classification', '')).font = INPUT_FONT
    ws2.cell(row=rr, column=5, value=s.get('anomaly_score', 0)).font = INPUT_FONT
    ws2.cell(row=rr, column=6, value=f"=(A{rr}+B{rr})/2/60000").font = FORMULA_FONT
    ws2.cell(row=rr, column=6).number_format = "0.00"
    ws2.cell(row=rr, column=7, value=f"=MAX(0,E{rr}-'Summary'!$C${tau_row})").font = FORMULA_FONT
    ws2.cell(row=rr, column=7).number_format = "0.000"
    if i == 0:
        ws2.cell(row=rr, column=8, value=0).font = FORMULA_FONT
    else:
        prev = rr - 1
        formula = f"=IF((F{rr}-F{prev})<=15,(F{rr}-F{prev})*(G{rr}+G{prev})/2,0)"
        ws2.cell(row=rr, column=8, value=formula).font = FORMULA_FONT
    ws2.cell(row=rr, column=8).number_format = "0.000"
    ws2.cell(row=rr, column=9, value=f'=TEXT(A{rr}/86400000+DATE(1970,1,1),"yyyy-mm-dd hh:mm")')
    ws2.cell(row=rr, column=9).font = FORMULA_FONT
    for cc in range(1, 10):
        ws2.cell(row=rr, column=cc).border = BORDER

last_seg_row = row0 + len(scored_segs) - 1

if len(scored_segs) > 1:
    chart2 = LineChart()
    chart2.title = f"Anomaly score per window (segmen)"
    chart2.y_axis.title = "Anomaly score"
    chart2.x_axis.title = "Window start (UTC)"
    chart2.height = 8
    chart2.width = 20
    data = Reference(ws2, min_col=5, min_row=1, max_row=last_seg_row)
    cats = Reference(ws2, min_col=9, min_row=2, max_row=last_seg_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws2.add_chart(chart2, f"A{last_seg_row+4}")

# ---------------- Raw HR sheet ----------------
ws3 = wb.create_sheet(f"RawHR")
ws3.sheet_view.showGridLines = False
headers3 = ["Timestamp (ms)", "Datetime (UTC)", "HR (bpm)", "RR (ms)", "RMSSD instant", "Activity", "Step count", "Label chart (text)"]
for c, h in enumerate(headers3, start=1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(headers3))
widths3 = [15, 20, 10, 10, 13, 12, 11, 18]
for c, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(c)].width = w

row0b = 2
for i, p in enumerate(polar):
    rr = row0b + i
    ws3.cell(row=rr, column=1, value=p.get('timestamp')).font = INPUT_FONT
    c2 = ws3.cell(row=rr, column=2, value=f"=A{rr}/86400000+DATE(1970,1,1)")
    c2.number_format = "yyyy-mm-dd hh:mm:ss"
    c2.font = FORMULA_FONT
    ws3.cell(row=rr, column=3, value=p.get('hr')).font = INPUT_FONT
    ws3.cell(row=rr, column=4, value=p.get('rr')).font = INPUT_FONT
    ws3.cell(row=rr, column=5, value=p.get('rrms')).font = INPUT_FONT
    ws3.cell(row=rr, column=6, value=p.get('activity')).font = INPUT_FONT
    ws3.cell(row=rr, column=7, value=p.get('step_count')).font = INPUT_FONT
    ws3.cell(row=rr, column=8, value=f'=TEXT(A{rr}/86400000+DATE(1970,1,1),"yyyy-mm-dd hh:mm:ss")')
    ws3.cell(row=rr, column=8).font = FORMULA_FONT
last_hr_row = row0b + len(polar) - 1
ws3.freeze_panes = "A2"

if len(polar) > 1:
    chart3 = LineChart()
    chart3.title = f"Grafik HR episode penuh"
    chart3.y_axis.title = "Heart Rate (bpm)"
    chart3.x_axis.title = "Waktu (UTC)"
    chart3.height = 10
    chart3.width = 26
    data3 = Reference(ws3, min_col=3, min_row=1, max_row=last_hr_row)
    chart3.add_data(data3, titles_from_data=True)
    cats3 = Reference(ws3, min_col=8, min_row=2, max_row=last_hr_row)
    chart3.set_categories(cats3)
    chart3.style = 2
    chart3.x_axis.tickLblSkip = max(1, len(polar) // 15)
    ws3.add_chart(chart3, "J2")

    info_col = 18 
    ws3.cell(row=2, column=info_col, value="Onset:").font = Font(name=FONT, bold=True, size=9)
    ws3.cell(row=2, column=info_col+1, value=ms_to_dt(onset).strftime("%Y-%m-%d %H:%M:%S")).font = NOTE_FONT
    ws3.cell(row=3, column=info_col, value="Peak:").font = Font(name=FONT, bold=True, size=9)
    ws3.cell(row=3, column=info_col+1, value=ms_to_dt(peak).strftime("%Y-%m-%d %H:%M:%S")).font = NOTE_FONT
    ws3.cell(row=4, column=info_col, value="Resolved:").font = Font(name=FONT, bold=True, size=9)
    ws3.cell(row=4, column=info_col+1, value=ms_to_dt(resolved).strftime("%Y-%m-%d %H:%M:%S")).font = NOTE_FONT

out_file = 'd:/Kerjaan/mern-vp/api/Analisis_Episode_TTR_AUCD.xlsx'
wb.save(out_file)
print(f"Saved to {out_file}")
